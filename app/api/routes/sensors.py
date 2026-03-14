"""
Sensors Router - Sensor data endpoints
"""
import io
from datetime import datetime, timezone, timedelta
from typing import Optional
from fastapi import APIRouter, HTTPException, Query, Depends
from fastapi.responses import StreamingResponse
from app.db.session import get_cursor
from app.utils import validate_sensor, RANGES
from app.api.deps import get_current_user

router = APIRouter(tags=["Sensors"])


@router.get("/latest/{sensor}")
def get_latest(sensor: str, current_user: dict = Depends(get_current_user)):
    """Mendapatkan data terbaru dari sensor"""
    table, columns = validate_sensor(sensor)

    try:
        with get_cursor() as cur:
            cur.execute(f"""
                SELECT * FROM {table}
                ORDER BY timestamp DESC
                LIMIT 1;
            """)
            row = cur.fetchone()

        if not row:
            return {"message": f"Belum ada data untuk sensor '{sensor}'"}

        return row

    except Exception as e:
        raise HTTPException(500, f"Database error: {e}")


@router.get("/history/{sensor}")
def get_history(
    sensor: str, 
    range: str,
    raw: Optional[bool] = Query(False, description="Jika True, ambil semua data tanpa sampling"),
    current_user: dict = Depends(get_current_user)
):
    """
    Mendapatkan history data sensor dengan rentang waktu tertentu.
    
    - **sensor**: nama sensor (dht22, mq2, pzem004t, bh1750)
    - **range**: rentang waktu (1h, 6h, 12h, 24h, 7d)
    - **raw**: jika True, ambil semua data tanpa sampling (hati-hati data besar)
    """
    table, columns = validate_sensor(sensor)

    if range not in RANGES:
        raise HTTPException(
            400,
            f"Rentang waktu tidak valid. Pilihan: {list(RANGES.keys())}"
        )

    wib = timezone(timedelta(hours=7))
    range_config = RANGES[range]
    time_limit = datetime.now(wib).replace(tzinfo=None) - range_config["delta"]
    interval = range_config["interval"]

    try:
        with get_cursor() as cur:
            if interval and not raw:
                # Query dengan sampling menggunakan time_bucket (TimescaleDB) atau date_trunc
                # Menggunakan pendekatan yang kompatibel dengan PostgreSQL biasa
                numeric_cols = [c for c in columns if c not in ["id", "timestamp"]]
                avg_cols = ", ".join([f"AVG({col}) as {col}" for col in numeric_cols])
                
                cur.execute(f"""
                    SELECT 
                        (to_timestamp(floor(extract(epoch from timestamp AT TIME ZONE 'Asia/Jakarta') / extract(epoch from interval '{interval}')) * extract(epoch from interval '{interval}')) AT TIME ZONE 'Asia/Jakarta')
                        AS time_bucket,
                        {avg_cols},
                        COUNT(*) as sample_count
                    FROM {table}
                    WHERE timestamp >= %s
                    GROUP BY time_bucket
                    ORDER BY time_bucket ASC;
                """, (time_limit,))
            else:
                # Query tanpa sampling (untuk 1h atau jika raw=True)
                cur.execute(f"""
                    SELECT *, timestamp AT TIME ZONE 'Asia/Jakarta' as timestamp_wib
                    FROM {table}
                    WHERE timestamp >= %s
                    ORDER BY timestamp ASC;
                """, (time_limit,))

            rows = cur.fetchall()

        return {
            "sensor": sensor,
            "range": range,
            "sampled": bool(interval and not raw),
            "interval": interval if (interval and not raw) else None,
            "count": len(rows),
            "data": rows
        }

    except Exception as e:
        raise HTTPException(500, f"Database error: {e}")


@router.get("/stats/{sensor}")
def get_stats(sensor: str, range: str, current_user: dict = Depends(get_current_user)):
    """
    Mendapatkan statistik agregasi dari sensor (min, max, avg).
    Sangat cepat karena hanya menghitung agregat.
    """
    table, columns = validate_sensor(sensor)

    if range not in RANGES:
        raise HTTPException(400, f"Rentang waktu tidak valid. Pilihan: {list(RANGES.keys())}")

    wib = timezone(timedelta(hours=7))
    range_config = RANGES[range]
    time_limit = datetime.now(wib).replace(tzinfo=None) - range_config["delta"]
    
    # Kolom numerik untuk agregasi
    numeric_cols = [c for c in columns if c not in ["id", "timestamp"]]
    
    # Build aggregation query
    agg_parts = []
    for col in numeric_cols:
        agg_parts.extend([
            f"MIN({col}) as {col}_min",
            f"MAX({col}) as {col}_max",
            f"AVG({col}) as {col}_avg"
        ])
    agg_query = ", ".join(agg_parts)

    try:
        with get_cursor() as cur:
            cur.execute(f"""
                SELECT 
                    COUNT(*) as total_records,
                    MIN(timestamp) as first_record,
                    MAX(timestamp) as last_record,
                    {agg_query}
                FROM {table}
                WHERE timestamp >= %s;
            """, (time_limit,))
            
            row = cur.fetchone()

        return {
            "sensor": sensor,
            "range": range,
            "stats": row
        }

    except Exception as e:
        raise HTTPException(500, f"Database error: {e}")


@router.get("/export/{sensor}")
def export_excel(sensor: str, current_user: dict = Depends(get_current_user)):
    """
    Download semua data history sensor dalam format Excel.
    """
    table, columns = validate_sensor(sensor)

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        
        wb = Workbook()
        ws = wb.active
        ws.title = f"{sensor.upper()} History"
        
        with get_cursor() as cur:
            # Fetch all data descending (newest first)
            col_list = ", ".join(columns)
            cur.execute(f"SELECT {col_list} FROM {table} ORDER BY timestamp DESC")
            
            # Use defined columns order
            col_names = columns
            
            # Style for header
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="0D9488", end_color="0D9488", fill_type="solid")
            
            # Write headers
            for col_idx, col_name in enumerate(col_names, 1):
                cell = ws.cell(row=1, column=col_idx, value=col_name.upper())
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")
            
            # Write data rows
            row_num = 2
            while True:
                rows = cur.fetchmany(1000)
                if not rows:
                    break
                for row in rows:
                    for col_idx, col_name in enumerate(col_names, 1):
                        value = row[col_name]
                        # Format timestamp for better readability
                        if col_name == 'timestamp' and value:
                            value = value.strftime('%Y-%m-%d %H:%M:%S') if hasattr(value, 'strftime') else str(value)
                        ws.cell(row=row_num, column=col_idx, value=value)
                    row_num += 1
            
            # Auto-adjust column widths
            for col_idx, col_name in enumerate(col_names, 1):
                ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = max(len(col_name) + 5, 15)
        
        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        filename = f"{sensor}_history_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )

    except ImportError:
        raise HTTPException(500, "openpyxl not installed. Run: pip install openpyxl")
    except Exception as e:
        raise HTTPException(500, f"Export error: {e}")


@router.delete("/{sensor}")
def delete_sensor_data(sensor: str, current_user: dict = Depends(get_current_user)):
    """
    Menghapus SEMUA data dari sensor tertentu.
    Tindakan ini tidak dapat dibatalkan.
    """
    table, columns = validate_sensor(sensor)

    # Optional: Check if user is admin (if role based access is needed)
    # if current_user.get("role") != "admin":
    #     raise HTTPException(403, "Hanya admin yang dapat menghapus data sensor")

    try:
        with get_cursor() as cur:
            # Use TRUNCATE for speed and complete removal
            cur.execute(f"TRUNCATE TABLE {table} RESTART IDENTITY;")
            
            # Log this action to history
            cur.execute(
                "INSERT INTO app_history (event, status, status_class) VALUES (%s, %s, %s)",
                (f"Data {sensor.upper()} dihapus oleh {current_user.get('username', 'user')}", "DELETED", "bg-red-100 text-red-800")
            )

        return {"message": f"Data sensor {sensor} berhasil dihapus permanen."}

    except Exception as e:
        raise HTTPException(500, f"Database error: {e}")
